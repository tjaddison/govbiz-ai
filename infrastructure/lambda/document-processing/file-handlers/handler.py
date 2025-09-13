"""
File Type Handlers Lambda Function for GovBizAI
Handles text extraction from various file formats:
- PDF (via PyMuPDF/Textract)
- Word documents (DOC, DOCX)
- Excel files (XLS, XLSX)
- Plain text files (TXT, CSV)
- HTML files
"""

import json
import logging
import boto3
from botocore.exceptions import ClientError
from typing import Dict, Any, List, Optional, Tuple
import os
import tempfile
import zipfile
from io import BytesIO
import csv
import mimetypes

# Document processing libraries
try:
    from docx import Document as DocxDocument
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import _Cell, Table
    from docx.text.paragraph import Paragraph
except ImportError:
    DocxDocument = None

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import magic
except ImportError:
    magic = None

from document_utils import (
    setup_logging, create_response, generate_correlation_id,
    validate_tenant_access, extract_metadata_from_s3_key,
    clean_extracted_text, get_file_type_from_extension,
    create_processing_metadata, sanitize_filename
)

# Initialize AWS clients
s3_client = boto3.client('s3')
lambda_client = boto3.client('lambda')

# Environment variables
PROCESSED_DOCUMENTS_BUCKET = os.environ['PROCESSED_DOCUMENTS_BUCKET']
TEXT_EXTRACTION_FUNCTION = os.environ.get('TEXT_EXTRACTION_FUNCTION', 'govbizai-text-extraction')

# Setup logging
setup_logging('file-handlers')
logger = logging.getLogger(__name__)


def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Main handler for file type processing

    Expected event structure:
    {
        "bucket": "bucket-name",
        "key": "file-key",
        "tenant_id": "tenant-uuid",
        "file_type": "auto" | "pdf" | "word" | "excel" | "text" | "html"
    }
    """
    correlation_id = generate_correlation_id()
    logger.info(f"Starting file type processing - Correlation ID: {correlation_id}")

    try:
        # Parse event parameters
        bucket = event['bucket']
        key = event['key']
        tenant_id = event['tenant_id']
        file_type = event.get('file_type', 'auto')

        logger.info(f"Processing file s3://{bucket}/{key} for tenant: {tenant_id}")

        # Auto-detect file type if needed
        if file_type == 'auto':
            file_type = detect_file_type(bucket, key)

        logger.info(f"Detected file type: {file_type}")

        # Route to appropriate handler
        if file_type == 'pdf':
            result = handle_pdf_file(bucket, key, tenant_id, correlation_id)
        elif file_type == 'word':
            result = handle_word_file(bucket, key, tenant_id, correlation_id)
        elif file_type == 'excel':
            result = handle_excel_file(bucket, key, tenant_id, correlation_id)
        elif file_type == 'text':
            result = handle_text_file(bucket, key, tenant_id, correlation_id)
        elif file_type == 'html':
            result = handle_html_file(bucket, key, tenant_id, correlation_id)
        elif file_type == 'csv':
            result = handle_csv_file(bucket, key, tenant_id, correlation_id)
        else:
            return create_response(400, {
                'message': f'Unsupported file type: {file_type}',
                'supported_types': ['pdf', 'word', 'excel', 'text', 'html', 'csv']
            }, correlation_id)

        logger.info(f"File processing completed successfully")

        return create_response(200, {
            'message': 'File processing completed successfully',
            'tenant_id': tenant_id,
            'file_type': file_type,
            'processing_result': result,
            'correlation_id': correlation_id
        }, correlation_id)

    except Exception as e:
        logger.error(f"Error in file processing: {str(e)}")
        return create_response(500, {
            'message': 'Internal server error',
            'error': str(e),
            'correlation_id': correlation_id
        }, correlation_id)


def detect_file_type(bucket: str, key: str) -> str:
    """
    Detect file type from extension and MIME type
    """
    # First try extension-based detection
    file_type = get_file_type_from_extension(key)

    if file_type != 'unknown':
        return file_type

    # Try MIME type detection if magic is available
    if magic:
        try:
            # Download first 2KB to check MIME type
            response = s3_client.get_object(
                Bucket=bucket,
                Key=key,
                Range='bytes=0-2047'
            )
            sample_data = response['Body'].read()

            mime_type = magic.from_buffer(sample_data, mime=True)

            mime_to_type = {
                'application/pdf': 'pdf',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'word',
                'application/msword': 'word',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'excel',
                'application/vnd.ms-excel': 'excel',
                'text/plain': 'text',
                'text/html': 'html',
                'text/csv': 'csv'
            }

            return mime_to_type.get(mime_type, 'text')

        except Exception as e:
            logger.warning(f"Could not detect MIME type: {str(e)}")

    # Default to text if detection fails
    return 'text'


def handle_pdf_file(bucket: str, key: str, tenant_id: str, correlation_id: str) -> Dict[str, Any]:
    """
    Handle PDF files by delegating to the text extraction Lambda
    """
    logger.info(f"Delegating PDF processing to text extraction function")

    try:
        # Invoke the text extraction Lambda function
        response = lambda_client.invoke(
            FunctionName=TEXT_EXTRACTION_FUNCTION,
            InvocationType='RequestResponse',
            Payload=json.dumps({
                'bucket': bucket,
                'key': key,
                'tenant_id': tenant_id
            })
        )

        result = json.loads(response['Payload'].read())

        if response['StatusCode'] == 200:
            return {
                'processing_method': 'delegated_to_text_extraction',
                'extraction_result': result
            }
        else:
            raise Exception(f"Text extraction function failed: {result}")

    except Exception as e:
        logger.error(f"Error delegating to text extraction function: {str(e)}")
        raise


def handle_word_file(bucket: str, key: str, tenant_id: str, correlation_id: str) -> Dict[str, Any]:
    """
    Handle Word documents (DOC, DOCX)
    """
    logger.info(f"Processing Word document: {key}")

    if not DocxDocument:
        raise Exception("python-docx library not available for Word document processing")

    try:
        # Download file
        with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
            s3_client.download_fileobj(bucket, key, tmp_file)
            tmp_file_path = tmp_file.name

        # Check if it's a DOCX file
        if key.lower().endswith('.docx'):
            extracted_data = extract_docx_content(tmp_file_path)
        else:
            # For DOC files, we'd need additional libraries like python-docx2txt
            # For now, return an error suggesting conversion
            os.unlink(tmp_file_path)
            raise Exception("Legacy DOC format not supported. Please convert to DOCX format.")

        os.unlink(tmp_file_path)

        # Process and store the extracted content
        return process_and_store_extracted_text(
            extracted_data,
            bucket,
            key,
            tenant_id,
            'word_extraction',
            correlation_id
        )

    except Exception as e:
        logger.error(f"Error processing Word document: {str(e)}")
        raise


def extract_docx_content(file_path: str) -> Dict[str, Any]:
    """
    Extract content from DOCX file
    """
    doc = DocxDocument(file_path)

    extracted_data = {
        'text_content': [],
        'tables': [],
        'metadata': {},
        'structure': []
    }

    # Extract core properties
    try:
        core_props = doc.core_properties
        extracted_data['metadata'] = {
            'title': core_props.title or '',
            'author': core_props.author or '',
            'subject': core_props.subject or '',
            'created': str(core_props.created) if core_props.created else '',
            'modified': str(core_props.modified) if core_props.modified else ''
        }
    except Exception:
        pass

    # Extract text content maintaining structure
    for element in doc.element.body:
        if isinstance(element, CT_P):
            # Paragraph
            paragraph = Paragraph(element, doc)
            text = paragraph.text.strip()
            if text:
                extracted_data['text_content'].append(text)
                extracted_data['structure'].append({
                    'type': 'paragraph',
                    'text': text,
                    'style': paragraph.style.name if paragraph.style else 'Normal'
                })

        elif isinstance(element, CT_Tbl):
            # Table
            table = Table(element, doc)
            table_data = extract_table_data(table)
            if table_data:
                extracted_data['tables'].append(table_data)
                extracted_data['structure'].append({
                    'type': 'table',
                    'rows': len(table_data),
                    'columns': len(table_data[0]) if table_data else 0
                })

    # Combine all text
    full_text = '\\n\\n'.join(extracted_data['text_content'])

    # Add table content to text
    for table in extracted_data['tables']:
        table_text = '\\n'.join([' | '.join(row) for row in table])
        full_text += f"\\n\\n[TABLE]\\n{table_text}\\n[/TABLE]"

    return {
        'full_text': full_text,
        'text_content': extracted_data['text_content'],
        'tables': extracted_data['tables'],
        'metadata': extracted_data['metadata'],
        'structure': extracted_data['structure'],
        'statistics': {
            'paragraphs': len(extracted_data['text_content']),
            'tables': len(extracted_data['tables']),
            'total_characters': len(full_text)
        }
    }


def extract_table_data(table: Table) -> List[List[str]]:
    """
    Extract data from a Word table
    """
    table_data = []

    for row in table.rows:
        row_data = []
        for cell in row.cells:
            cell_text = cell.text.strip()
            row_data.append(cell_text)
        table_data.append(row_data)

    return table_data


def handle_excel_file(bucket: str, key: str, tenant_id: str, correlation_id: str) -> Dict[str, Any]:
    """
    Handle Excel files (XLS, XLSX)
    """
    logger.info(f"Processing Excel file: {key}")

    try:
        # Download file
        with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
            s3_client.download_fileobj(bucket, key, tmp_file)
            tmp_file_path = tmp_file.name

        if key.lower().endswith('.xlsx'):
            if not openpyxl:
                raise Exception("openpyxl library not available for XLSX processing")
            extracted_data = extract_xlsx_content(tmp_file_path)
        else:
            if not xlrd:
                raise Exception("xlrd library not available for XLS processing")
            extracted_data = extract_xls_content(tmp_file_path)

        os.unlink(tmp_file_path)

        # Process and store the extracted content
        return process_and_store_extracted_text(
            extracted_data,
            bucket,
            key,
            tenant_id,
            'excel_extraction',
            correlation_id
        )

    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise


def extract_xlsx_content(file_path: str) -> Dict[str, Any]:
    """
    Extract content from XLSX file
    """
    workbook = load_workbook(file_path, data_only=True)

    extracted_data = {
        'worksheets': {},
        'metadata': {
            'worksheet_names': workbook.sheetnames,
            'total_worksheets': len(workbook.sheetnames)
        }
    }

    full_text_parts = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # Extract data from worksheet
        sheet_data = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else '')
            sheet_data.append(row_data)

        extracted_data['worksheets'][sheet_name] = {
            'data': sheet_data,
            'dimensions': f"{max_row}x{max_col}"
        }

        # Convert to text
        sheet_text = f"\\n\\n=== {sheet_name} ===\\n"
        for row in sheet_data:
            if any(cell.strip() for cell in row):  # Skip empty rows
                sheet_text += ' | '.join(row) + '\\n'

        full_text_parts.append(sheet_text)

    return {
        'full_text': '\\n'.join(full_text_parts),
        'worksheets': extracted_data['worksheets'],
        'metadata': extracted_data['metadata'],
        'statistics': {
            'total_worksheets': len(workbook.sheetnames),
            'total_characters': len('\\n'.join(full_text_parts))
        }
    }


def extract_xls_content(file_path: str) -> Dict[str, Any]:
    """
    Extract content from XLS file
    """
    workbook = xlrd.open_workbook(file_path)

    extracted_data = {
        'worksheets': {},
        'metadata': {
            'worksheet_names': workbook.sheet_names(),
            'total_worksheets': workbook.nsheets
        }
    }

    full_text_parts = []

    for sheet_name in workbook.sheet_names():
        worksheet = workbook.sheet_by_name(sheet_name)

        # Extract data from worksheet
        sheet_data = []
        for row_idx in range(worksheet.nrows):
            row_data = []
            for col_idx in range(worksheet.ncols):
                cell_value = worksheet.cell_value(row_idx, col_idx)
                row_data.append(str(cell_value) if cell_value else '')
            sheet_data.append(row_data)

        extracted_data['worksheets'][sheet_name] = {
            'data': sheet_data,
            'dimensions': f"{worksheet.nrows}x{worksheet.ncols}"
        }

        # Convert to text
        sheet_text = f"\\n\\n=== {sheet_name} ===\\n"
        for row in sheet_data:
            if any(cell.strip() for cell in row):  # Skip empty rows
                sheet_text += ' | '.join(row) + '\\n'

        full_text_parts.append(sheet_text)

    return {
        'full_text': '\\n'.join(full_text_parts),
        'worksheets': extracted_data['worksheets'],
        'metadata': extracted_data['metadata'],
        'statistics': {
            'total_worksheets': workbook.nsheets,
            'total_characters': len('\\n'.join(full_text_parts))
        }
    }


def handle_text_file(bucket: str, key: str, tenant_id: str, correlation_id: str) -> Dict[str, Any]:
    """
    Handle plain text files
    """
    logger.info(f"Processing text file: {key}")

    try:
        # Download and read file
        response = s3_client.get_object(Bucket=bucket, Key=key)
        content = response['Body'].read()

        # Try to decode with different encodings
        text = decode_text_content(content)

        extracted_data = {
            'full_text': text,
            'metadata': {
                'encoding': 'utf-8',  # Simplified
                'file_size': len(content)
            },
            'statistics': {
                'total_characters': len(text),
                'total_lines': len(text.split('\\n')),
                'total_words': len(text.split())
            }
        }

        # Process and store the extracted content
        return process_and_store_extracted_text(
            extracted_data,
            bucket,
            key,
            tenant_id,
            'text_extraction',
            correlation_id
        )

    except Exception as e:
        logger.error(f"Error processing text file: {str(e)}")
        raise


def handle_html_file(bucket: str, key: str, tenant_id: str, correlation_id: str) -> Dict[str, Any]:
    """
    Handle HTML files
    """
    logger.info(f"Processing HTML file: {key}")

    if not BeautifulSoup:
        raise Exception("BeautifulSoup library not available for HTML processing")

    try:
        # Download and read file
        response = s3_client.get_object(Bucket=bucket, Key=key)
        content = response['Body'].read()

        # Try to decode
        html_content = decode_text_content(content)

        # Parse HTML
        soup = BeautifulSoup(html_content, 'html.parser')

        # Extract text content
        text = soup.get_text(separator='\\n', strip=True)

        # Extract metadata
        metadata = {}
        if soup.title:
            metadata['title'] = soup.title.string

        # Extract headings for structure
        headings = []
        for i in range(1, 7):
            for heading in soup.find_all(f'h{i}'):
                headings.append({
                    'level': i,
                    'text': heading.get_text(strip=True)
                })

        extracted_data = {
            'full_text': text,
            'metadata': metadata,
            'headings': headings,
            'statistics': {
                'total_characters': len(text),
                'total_headings': len(headings)
            }
        }

        # Process and store the extracted content
        return process_and_store_extracted_text(
            extracted_data,
            bucket,
            key,
            tenant_id,
            'html_extraction',
            correlation_id
        )

    except Exception as e:
        logger.error(f"Error processing HTML file: {str(e)}")
        raise


def handle_csv_file(bucket: str, key: str, tenant_id: str, correlation_id: str) -> Dict[str, Any]:
    """
    Handle CSV files
    """
    logger.info(f"Processing CSV file: {key}")

    try:
        # Download and read file
        response = s3_client.get_object(Bucket=bucket, Key=key)
        content = response['Body'].read()

        # Try to decode
        text_content = decode_text_content(content)

        # Parse CSV
        csv_reader = csv.reader(text_content.splitlines())
        rows = list(csv_reader)

        # Convert to structured format
        if rows:
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []

            # Convert to text
            text_parts = []
            text_parts.append("=== CSV Data ===")
            text_parts.append(" | ".join(headers))
            text_parts.append("-" * 40)

            for row in data_rows[:100]:  # Limit to first 100 rows
                text_parts.append(" | ".join(row))

            if len(data_rows) > 100:
                text_parts.append(f"... and {len(data_rows) - 100} more rows")

            full_text = "\\n".join(text_parts)
        else:
            full_text = "Empty CSV file"

        extracted_data = {
            'full_text': full_text,
            'csv_data': {
                'headers': headers if rows else [],
                'row_count': len(data_rows) if rows else 0,
                'column_count': len(headers) if rows else 0
            },
            'metadata': {
                'total_rows': len(rows),
                'has_headers': bool(rows)
            },
            'statistics': {
                'total_characters': len(full_text),
                'total_rows': len(rows)
            }
        }

        # Process and store the extracted content
        return process_and_store_extracted_text(
            extracted_data,
            bucket,
            key,
            tenant_id,
            'csv_extraction',
            correlation_id
        )

    except Exception as e:
        logger.error(f"Error processing CSV file: {str(e)}")
        raise


def decode_text_content(content: bytes) -> str:
    """
    Try to decode text content with different encodings
    """
    encodings = ['utf-8', 'utf-16', 'latin-1', 'cp1252']

    for encoding in encodings:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue

    # If all fail, decode with errors ignored
    return content.decode('utf-8', errors='ignore')


def process_and_store_extracted_text(
    extracted_data: Dict[str, Any],
    source_bucket: str,
    source_key: str,
    tenant_id: str,
    processing_type: str,
    correlation_id: str
) -> Dict[str, Any]:
    """
    Process extracted text and store results
    """
    try:
        # Clean the extracted text
        cleaned_text = clean_extracted_text(extracted_data['full_text'])

        # Create processed document data
        processed_data = {
            'tenant_id': tenant_id,
            'source_bucket': source_bucket,
            'source_key': source_key,
            'processing_type': processing_type,
            'correlation_id': correlation_id,
            'full_text': cleaned_text,
            'original_data': extracted_data,
            'statistics': {
                'original_characters': len(extracted_data['full_text']),
                'cleaned_characters': len(cleaned_text),
                'reduction_percentage': (
                    (len(extracted_data['full_text']) - len(cleaned_text)) /
                    max(len(extracted_data['full_text']), 1)
                ) * 100
            },
            'processing_metadata': create_processing_metadata(
                {
                    'bucket': source_bucket,
                    'key': source_key,
                    'type': processing_type.replace('_extraction', '')
                },
                processing_type,
                tenant_id,
                {
                    'correlation_id': correlation_id
                }
            )
        }

        # Store processed results in S3
        base_name = os.path.splitext(os.path.basename(source_key))[0]
        output_key = f"tenants/{tenant_id}/processed/{base_name}_extracted.json"

        s3_client.put_object(
            Bucket=PROCESSED_DOCUMENTS_BUCKET,
            Key=output_key,
            Body=json.dumps(processed_data, default=str, indent=2),
            ContentType='application/json',
            ServerSideEncryption='AES256',
            Metadata={
                'tenant-id': tenant_id,
                'processing-type': processing_type,
                'correlation-id': correlation_id,
                'source-key': source_key
            }
        )

        logger.info(f"Processed document stored at s3://{PROCESSED_DOCUMENTS_BUCKET}/{output_key}")

        return {
            'output_location': f"s3://{PROCESSED_DOCUMENTS_BUCKET}/{output_key}",
            'characters_extracted': len(cleaned_text),
            'processing_method': processing_type
        }

    except Exception as e:
        logger.error(f"Error storing processed document: {str(e)}")
        raise